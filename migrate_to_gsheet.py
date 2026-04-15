"""One-off migration: 排班.xlsx -> Google Sheet.

Uploads every sheet from the local workbook to the target Google Sheet,
preserving values and holiday yellow background (FFEB9C).
"""
import openpyxl
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials

XLSX_PATH = '排班.xlsx'
SHEET_ID = '10ilVOmJrr8jjfnMMbtj60tAIIAe1YX3ZRU1RLgn6Elk'
CREDS_PATH = '.gsa.json'
YELLOW_RGB = 'FFEB9C'

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
]


def fmt_value(v):
    """Convert openpyxl cell value to a clean string/number for Google Sheets."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def extract_sheet(ws):
    """Return (values_2d, yellow_cells_list[(row, col)], rows, cols)."""
    max_r, max_c = 0, 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                max_r = max(max_r, c.row)
                max_c = max(max_c, c.column)
    if max_r == 0:
        return [], [], 1, 1
    values = []
    yellow = []
    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        row_vals = []
        for c in row:
            row_vals.append(fmt_value(c.value))
            if c.fill and c.fill.fgColor and c.fill.fgColor.rgb:
                rgb = str(c.fill.fgColor.rgb).upper()
                if YELLOW_RGB in rgb:
                    yellow.append((c.row, c.column))
        values.append(row_vals)
    return values, yellow, max_r, max_c


def main():
    creds = Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SHEET_ID)
    print(f'Opened: {sh.title}')

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    existing_titles = [ws.title for ws in sh.worksheets()]

    for name in wb.sheetnames:
        src = wb[name]
        values, yellow, rows, cols = extract_sheet(src)
        print(f'\n[{name}] {rows}x{cols}, {len(yellow)} yellow cells')

        if name in existing_titles:
            dst = sh.worksheet(name)
            dst.clear()
            dst.resize(rows=max(rows, 10), cols=max(cols, 5))
        else:
            dst = sh.add_worksheet(title=name, rows=max(rows, 10), cols=max(cols, 5))

        if values:
            dst.update(range_name='A1', values=values, value_input_option='USER_ENTERED')

        if yellow:
            sheet_id = dst.id
            requests = []
            yellow_color = {'red': 1.0, 'green': 235 / 255, 'blue': 156 / 255}
            for (r, c) in yellow:
                requests.append({
                    'repeatCell': {
                        'range': {
                            'sheetId': sheet_id,
                            'startRowIndex': r - 1,
                            'endRowIndex': r,
                            'startColumnIndex': c - 1,
                            'endColumnIndex': c,
                        },
                        'cell': {
                            'userEnteredFormat': {
                                'backgroundColor': yellow_color,
                            }
                        },
                        'fields': 'userEnteredFormat.backgroundColor',
                    }
                })
            if requests:
                sh.batch_update({'requests': requests})
                print(f'  -> applied {len(requests)} yellow cells')

    final_titles = [ws.title for ws in sh.worksheets()]
    if '工作表1' in final_titles and len(final_titles) > 1:
        default_ws = sh.worksheet('工作表1')
        sh.del_worksheet(default_ws)
        print('\nDeleted default worksheet 工作表1')

    print('\nMigration complete.')
    print('Final sheets:')
    for ws in sh.worksheets():
        print(f'  - {ws.title}')


if __name__ == '__main__':
    main()
