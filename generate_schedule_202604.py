import pandas as pd
from datetime import date, timedelta
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

# --- Configuration ---
year, month = 2026, 4
sheet_name = "202604"
excel_file = '排班.xlsx'

# Doctors
crs = ["麒翔", "見賢", "常胤"]
vs_list = ["廖瑀", "昭佑", "朝允", "則瑋"]
inter_mid = ["展瀚", "建寬"]

# Holidays in April 2026
holidays = [date(2026, 4, 3), date(2026, 4, 4), date(2026, 4, 5), date(2026, 4, 6)]

def is_holiday(d):
    return d.weekday() >= 5 or d in holidays

# New Statistical Class Logic (連假邏輯)
def get_stat_type(d):
    # 4/2 is "Friday" because 4/3-6 is a holiday
    if d == date(2026, 4, 2): return "週五班"
    # 4/3-5 are "Saturday" because they are holidays but not the last day
    if d in [date(2026, 4, 3), date(2026, 4, 4), date(2026, 4, 5)]: return "週六班"
    # 4/6 is "Sunday" because it's the last day of long weekend
    if d == date(2026, 4, 6): return "週日班"
    
    # Regular rules
    if is_holiday(d):
        if d.weekday() == 5: return "週六班"
        if d.weekday() == 6: return "週日班"
        return "假日其他" 
    else:
        if d.weekday() == 4: return "週五班"
        return "平日"

# Fixed Preferences
fixed = {
    date(2026, 4, 11): "朝允",
    date(2026, 4, 18): "昭佑",
    date(2026, 4, 25): "廖瑀",
    date(2026, 4, 26): "則瑋",
    date(2026, 4, 1): "展瀚",
    date(2026, 4, 8): "展瀚",
    date(2026, 4, 21): "展瀚",
    date(2026, 4, 22): "展瀚",
}

avoid = {
    "見賢": [date(2026, 4, d) for d in range(18, 27)],
    "常胤": [date(2026, 4, 15)],
}

# --- Solver ---
def solve():
    days = [date(year, month, d) for d in range(1, 31)]
    schedule = {}
    cr_w_counts = {n: 0 for n in crs}
    cr_h_counts = {n: 0 for n in crs}
    jk_count = 0
    
    def backtrack(d_idx):
        nonlocal jk_count
        if d_idx == 30: return True
        d = days[d_idx]
        
        candidates = []
        if d in fixed:
            candidates = [fixed[d]]
        elif is_holiday(d):
            candidates = crs
        else:
            candidates = crs + ["建寬"]
            
        # Balance counts
        candidates.sort(key=lambda x: (cr_w_counts.get(x, 0) if not is_holiday(d) else cr_h_counts.get(x, 0)))

        for name in candidates:
            # Rule: No back-to-back (Except Zhanhan)
            if name != "展瀚" and d_idx > 0 and schedule.get(days[d_idx-1]) == name: continue
            # Avoid
            if name in avoid and d in avoid[name]: continue
            # Max counts
            if name in crs:
                if is_holiday(d) and cr_h_counts[name] >= 2: continue
                if not is_holiday(d) and cr_w_counts[name] >= 5: continue
            if name == "建寬" and jk_count >= 1: continue
            
            schedule[d] = name
            if name in crs:
                if is_holiday(d): cr_h_counts[name] += 1
                else: cr_w_counts[name] += 1
            if name == "建寬": jk_count += 1
            
            if backtrack(d_idx + 1): return True
            
            # Backtrack
            if name in crs:
                if is_holiday(d): cr_h_counts[name] -= 1
                else: cr_w_counts[name] -= 1
            if name == "建寬": jk_count -= 1
            del schedule[d]
        return False

    if backtrack(0): return schedule
    return None

result = solve()

# --- Export & Formatting ---
try:
    wb = load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()

if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

fill_holiday = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Header
days_header = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
for col, day in enumerate(days_header, 1):
    cell = ws.cell(row=1, column=col, value=day)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Calendar Filling
month_cal = calendar.monthcalendar(year, month)
for r_idx, week in enumerate(month_cal):
    for c_idx, day in enumerate(week, 1):
        if day == 0: continue
        d_obj = date(year, month, day)
        cell_d = ws.cell(row=r_idx*2+2, column=c_idx, value=day)
        cell_n = ws.cell(row=r_idx*2+3, column=c_idx, value=result.get(d_obj, ""))
        cell_d.border = cell_n.border = thin_border
        cell_n.font = Font(bold=True)
        cell_n.alignment = Alignment(horizontal='center', vertical='center')
        if is_holiday(d_obj):
            cell_d.fill = cell_n.fill = fill_holiday

for i in range(1, 8): ws.column_dimensions[chr(64+i)].width = 18

# Stats
stats = []
for name in crs + vs_list + inter_mid:
    personal = [d for d, n in result.items() if n == name]
    stats.append({
        "姓名": name,
        "平日班": len([d for d in personal if not is_holiday(d)]),
        "假日班": len([d for d in personal if is_holiday(d)]),
        "週五班": len([d for d in personal if get_stat_type(d) == "週五班"]),
        "週六班": len([d for d in personal if get_stat_type(d) == "週六班"]),
        "週日班": len([d for d in personal if get_stat_type(d) == "週日班"])
    })
df_stats = pd.DataFrame(stats)

# Reset and Add to baseline
baseline = {
    "見賢": {"平日": 31, "週五": 10, "週六": 4, "週日": 10, "假日": 15},
    "麒翔": {"平日": 34, "週五": 7, "週六": 7, "週日": 7, "假日": 16},
    "常胤": {"平日": 31, "週五": 9, "週六": 6, "週日": 9, "假日": 17},
    "廖瑀": {"平日": 0, "週五": 3, "週六": 8, "週日": 0, "假日": 8},
    "則瑋": {"平日": 4, "週五": 1, "週六": 1, "週日": 3, "假日": 5},
    "昭佑": {"平日": 0, "週五": 1, "週六": 2, "週日": 4, "假日": 6},
    "朝允": {"平日": 0, "週五": 2, "週六": 6, "週日": 1, "假日": 8}
}

if "值班總數統計" in wb.sheetnames:
    ws_total = wb["值班總數統計"]
    header_map = {ws_total.cell(1, c).value: c for c in range(1, ws_total.max_column + 1)}
    for name, base in baseline.items():
        s_row_list = df_stats[df_stats["姓名"] == name]
        if s_row_list.empty: continue
        s_row = s_row_list.iloc[0]
        for r in range(2, ws_total.max_row + 1):
            if ws_total.cell(r, header_map["姓名"]).value == name:
                ws_total.cell(r, header_map["平日班 (一-四)"], value=base["平日"] + s_row["平日班"])
                ws_total.cell(r, header_map["周五班"], value=base["週五"] + s_row["週五班"])
                ws_total.cell(r, header_map["周六班"], value=base["週六"] + s_row["週六班"])
                ws_total.cell(r, header_map["周日班"], value=base["週日"] + s_row["週日班"])
                ws_total.cell(r, header_map["假日班 (含六日及國定假日)"], value=base["假日"] + s_row["假日班"])
                break

wb.save(excel_file)
print("Success! Long-weekend stat logic applied.")
